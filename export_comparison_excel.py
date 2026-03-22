from pathlib import Path
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


base = Path('/Users/tegin/Desktop/UCCR/raw_data')
out_path = base / 'processed' / 'UCCR_Moya_Comparison.xlsx'
analysis_results_path = base / 'processed' / 'analysis_2ug_results.json'
urine_results_path = base / 'processed' / 'urine_cortisol_analysis_results.json'
roc_summary_path = base / 'processed' / 'roc_scenarios' / 'roc_scenarios_summary.csv'


def pct_str(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return 'NA'
    return f'{float(value) * 100:.1f}%'


def cutoff_str(best: dict | None) -> str:
    if not isinstance(best, dict):
        return 'NA'
    direction = str(best.get('direction', 'lower'))
    op = '<=' if direction == 'lower' else '>='
    threshold = best.get('threshold')
    if threshold is None or pd.isna(threshold):
        return 'NA'
    return f'{op}{float(threshold):.2f}'


def build_additional_conditions_table(results_path: Path, analysis_label: str) -> pd.DataFrame:
    columns = [
        'Analysis',
        'Scenario',
        'Outcome',
        'Urine test type',
        'Optimal cutpoint',
        'Sensitivity',
        'Specificity',
        'PPV',
        'NPV',
        'Accuracy',
        'n (pos/neg)',
        'Excluded rows',
        'Remaining rows',
    ]

    payload = json.loads(results_path.read_text())
    additional = payload.get('additional_conditions', {})
    rows = []

    condition_order = [
        ('exclude_missing_post_if_basal_lt1ugdl', 'post missing AND basal <1 ug/dL'),
        ('exclude_missing_post_if_basal_lt2ugdl', 'post missing AND basal <2 ug/dL'),
    ]
    outcome_order = [
        ('best_cutoffs_diag_by_testtype', 'HA diagnosis (post-ACTH <=2 ug/dL)'),
        ('best_cutoffs_excl_by_testtype', 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)'),
    ]

    for condition_key, scenario_label in condition_order:
        condition_data = additional.get(condition_key, {})
        excluded_rows = int(condition_data.get('excluded_rows', 0))
        remaining_rows = int(condition_data.get('overall_counts', {}).get('n_total', 0))

        for outcome_key, outcome_label in outcome_order:
            outcome_map = condition_data.get(outcome_key, {})
            for urine_type in ('CLIApost', 'RIA'):
                best = outcome_map.get(urine_type)
                if best is None:
                    continue
                rows.append(
                    {
                        'Analysis': analysis_label,
                        'Scenario': scenario_label,
                        'Outcome': outcome_label,
                        'Urine test type': urine_type,
                        'Optimal cutpoint': cutoff_str(best),
                        'Sensitivity': pct_str(best.get('sensitivity')),
                        'Specificity': pct_str(best.get('specificity')),
                        'PPV': pct_str(best.get('ppv')),
                        'NPV': pct_str(best.get('npv')),
                        'Accuracy': pct_str(best.get('accuracy')),
                        'n (pos/neg)': f"{int(best.get('n_total', 0))} ({int(best.get('n_pos', 0))}/{int(best.get('n_neg', 0))})",
                        'Excluded rows': excluded_rows,
                        'Remaining rows': remaining_rows,
                    }
                )

    return pd.DataFrame(rows, columns=columns)


def build_roc_scenarios_table(summary_path: Path) -> pd.DataFrame:
    columns = [
        'Scenario',
        'Analysis',
        'Outcome',
        'Urine test type',
        'AUC',
        'Best cutoff',
        'Sensitivity',
        'Specificity',
        'Accuracy',
        'n (pos/neg)',
        'Excluded rows',
        'Remaining rows',
        'ROC figure path',
    ]

    if not summary_path.exists():
        return pd.DataFrame(columns=columns)

    raw = pd.read_csv(summary_path)
    rows = []
    for _, r in raw.iterrows():
        op = '<=' if str(r.get('direction', 'lower')) == 'lower' else '>='
        rows.append(
            {
                'Scenario': r.get('scenario_label', ''),
                'Analysis': r.get('analysis', ''),
                'Outcome': r.get('outcome', ''),
                'Urine test type': r.get('urine_test_type', ''),
                'AUC': f"{float(r.get('auc', float('nan'))):.3f}",
                'Best cutoff': f"{op}{float(r.get('best_threshold', float('nan'))):.2f}",
                'Sensitivity': pct_str(r.get('best_sensitivity')),
                'Specificity': pct_str(r.get('best_specificity')),
                'Accuracy': pct_str(r.get('best_accuracy')),
                'n (pos/neg)': f"{int(r.get('n_total_model', 0))} ({int(r.get('n_pos', 0))}/{int(r.get('n_neg', 0))})",
                'Excluded rows': int(r.get('excluded_rows', 0)),
                'Remaining rows': int(r.get('remaining_rows', 0)),
                'ROC figure path': r.get('roc_png', ''),
            }
        )

    table = pd.DataFrame(rows, columns=columns)
    return table.sort_values(['Scenario', 'Analysis', 'Outcome', 'Urine test type'])


def build_overall_interpretation_table(
    table_2: pd.DataFrame,
    table_5: pd.DataFrame,
    table_6: pd.DataFrame,
    table_8: pd.DataFrame,
    table_9: pd.DataFrame,
    table_11: pd.DataFrame,
    table_12: pd.DataFrame,
    table_13: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    cohort = table_2.loc[
        table_2['Parameter'] == 'Number of dogs used',
        'Current Dataset (This Workspace)',
    ].iloc[0]
    rows.append(
        {
            'Topic': 'Cohort and analysis definitions',
            'Key finding': str(cohort),
            'Interpretation': 'The current dataset is substantially larger than Moya and supports stratified CLIApost/RIA analyses.',
        }
    )

    uccr_diag_clia = table_5.loc[table_5['Assay/Protocol'] == 'Current CLIApost'].iloc[0]
    uccr_diag_ria = table_5.loc[table_5['Assay/Protocol'] == 'Current RIA'].iloc[0]
    rows.append(
        {
            'Topic': 'UCCR diagnosis performance',
            'Key finding': (
                f"CLIApost {uccr_diag_clia['Optimal cutpoint']} (sens {uccr_diag_clia['Sensitivity']}, spec {uccr_diag_clia['Specificity']}); "
                f"RIA {uccr_diag_ria['Optimal cutpoint']} (sens {uccr_diag_ria['Sensitivity']}, spec {uccr_diag_ria['Specificity']})"
            ),
            'Interpretation': 'UCCR diagnosis performance is strong in both assays, with higher specificity in RIA and higher sensitivity in CLIApost.',
        }
    )

    uccr_excl_clia = table_6.loc[table_6['Assay/Protocol'] == 'Current CLIApost'].iloc[0]
    uccr_excl_ria = table_6.loc[table_6['Assay/Protocol'] == 'Current RIA'].iloc[0]
    rows.append(
        {
            'Topic': 'UCCR exclusion performance',
            'Key finding': (
                f"CLIApost {uccr_excl_clia['Optimal cutpoint']} (sens {uccr_excl_clia['Sensitivity']}, spec {uccr_excl_clia['Specificity']}); "
                f"RIA {uccr_excl_ria['Optimal cutpoint']} (sens {uccr_excl_ria['Sensitivity']}, spec {uccr_excl_ria['Specificity']})"
            ),
            'Interpretation': 'UCCR supports high sensitivity exclusion in both assays, with stronger specificity in RIA.',
        }
    )

    urine_diag_clia = table_8.loc[table_8['Assay/Protocol'] == 'CLIApost'].iloc[0]
    urine_diag_ria = table_8.loc[table_8['Assay/Protocol'] == 'RIA'].iloc[0]
    rows.append(
        {
            'Topic': 'Urine cortisol diagnosis performance',
            'Key finding': (
                f"CLIApost {urine_diag_clia['Optimal cutpoint (nmol/L)']} (sens {urine_diag_clia['Sensitivity']}, spec {urine_diag_clia['Specificity']}); "
                f"RIA {urine_diag_ria['Optimal cutpoint (nmol/L)']} (sens {urine_diag_ria['Sensitivity']}, spec {urine_diag_ria['Specificity']})"
            ),
            'Interpretation': 'Urine cortisol diagnosis is excellent, with perfect sensitivity in RIA at the selected threshold.',
        }
    )

    urine_excl_clia = table_9.loc[table_9['Assay/Protocol'] == 'CLIApost'].iloc[0]
    urine_excl_ria = table_9.loc[table_9['Assay/Protocol'] == 'RIA'].iloc[0]
    rows.append(
        {
            'Topic': 'Urine cortisol exclusion performance',
            'Key finding': (
                f"CLIApost {urine_excl_clia['Optimal cutpoint (nmol/L)']} (sens {urine_excl_clia['Sensitivity']}, spec {urine_excl_clia['Specificity']}); "
                f"RIA {urine_excl_ria['Optimal cutpoint (nmol/L)']} (sens {urine_excl_ria['Sensitivity']}, spec {urine_excl_ria['Specificity']})"
            ),
            'Interpretation': 'Urine cortisol exclusion is especially strong in RIA where specificity reaches 100.0% at the selected cutoff.',
        }
    )

    addl_uccr_clia_lt1 = table_11.loc[
        (table_11['Analysis'] == 'UCCR')
        & (table_11['Scenario'] == 'post missing AND basal <1 ug/dL')
        & (table_11['Outcome'] == 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)')
        & (table_11['Urine test type'] == 'CLIApost')
    ].iloc[0]
    addl_uccr_clia_lt2 = table_11.loc[
        (table_11['Analysis'] == 'UCCR')
        & (table_11['Scenario'] == 'post missing AND basal <2 ug/dL')
        & (table_11['Outcome'] == 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)')
        & (table_11['Urine test type'] == 'CLIApost')
    ].iloc[0]
    rows.append(
        {
            'Topic': 'Effect of additional-condition exclusions',
            'Key finding': (
                f"In UCCR CLIApost exclusion, specificity improved from {uccr_excl_clia['Specificity']} (base) "
                f"to {addl_uccr_clia_lt1['Specificity']} with <1 filtering and {addl_uccr_clia_lt2['Specificity']} with <2 filtering"
            ),
            'Interpretation': 'Removing missing-post low-basal cases mainly strengthens CLIApost exclusion specificity while preserving sensitivity.',
        }
    )

    if not table_13.empty:
        auc_vals = pd.to_numeric(table_13['AUC'], errors='coerce').dropna()
        auc_min = f"{auc_vals.min():.3f}" if len(auc_vals) else 'NA'
        auc_max = f"{auc_vals.max():.3f}" if len(auc_vals) else 'NA'
    else:
        auc_min = 'NA'
        auc_max = 'NA'
    rows.append(
        {
            'Topic': 'ROC discrimination across all scenarios',
            'Key finding': f'AUC range {auc_min} to {auc_max} across 24 curves',
            'Interpretation': 'All scenario-specific models show high discrimination, indicating robust assay-based classification performance.',
        }
    )

    rows.append(
        {
            'Topic': 'Perfect-classification cutoffs',
            'Key finding': 'No group achieved 100% accuracy; several achieved isolated 100% sensitivity or 100% specificity cutoffs',
            'Interpretation': 'The assays support strong but not perfect classification, so threshold choice should reflect whether sensitivity or specificity is prioritized.',
        }
    )

    return pd.DataFrame(rows, columns=['Topic', 'Key finding', 'Interpretation'])


table_2 = pd.DataFrame(
    [
        {
            'Parameter': 'Number of dogs used',
            'Moya et al. (2022)': '148 total (41 HA, 107 NAI)',
            'Current Dataset (This Workspace)': '330 total merged rows; 306 usable for HA diagnosis analysis (44 HA, 262 non-HA)',
        },
        {
            'Parameter': 'Inclusion criteria',
            'Moya et al. (2022)': 'Dogs tested for HA via basal cortisol or ACTH stimulation; no age/breed/sex restrictions',
            'Current Dataset (This Workspace)': 'Merged from 3 workbooks; UCCR_rawdata restricted to Final Combined Group; data-quality filters and deduplication applied',
        },
        {
            'Parameter': 'Exclusion criteria',
            'Moya et al. (2022)': 'Hyperadrenocorticism; recent glucocorticoids/azole antifungals; equivocal status',
            'Current Dataset (This Workspace)': 'Medication-history exclusions not available in source files; only data-quality exclusions applied',
        },
        {
            'Parameter': 'Serum cutoff to diagnose HA',
            'Moya et al. (2022)': 'Pre and post ACTH cortisol <=55 nmol/L (2 ug/dL)',
            'Current Dataset (This Workspace)': 'Primary diagnostic analysis used post-ACTH cortisol <=55.18 nmol/L (2 ug/dL)',
        },
        {
            'Parameter': 'Serum cutoff to exclude HA',
            'Moya et al. (2022)': 'Basal >55 nmol/L OR post-ACTH >138 nmol/L',
            'Current Dataset (This Workspace)': 'Comparison run used basal >55.18 nmol/L OR post-ACTH >55.18 nmol/L (2 ug/dL)',
        },
    ]
)


table_3 = pd.DataFrame(
    [
        {'Assay/Protocol': 'Moya RIA', 'HA median (IQR)': '0 (0 to 0)', 'Non-HA median (IQR)': '22 (7 to 47.3)'},
        {'Assay/Protocol': 'Current RIA', 'HA median (IQR)': '0.0 (0.0 to 1.0)', 'Non-HA median (IQR)': '22.0 (13.0 to 42.5)'},
        {'Assay/Protocol': 'Moya CLIA (=CLIApre)', 'HA median (IQR)': '1 (0 to 1.3)', 'Non-HA median (IQR)': '71 (30 to 127)'},
        {'Assay/Protocol': 'Current CLIApost', 'HA median (IQR)': '0.5 (0.0 to 1.8225)', 'Non-HA median (IQR)': '19.88 (10.615 to 45.92)'},
    ]
)


table_4 = pd.DataFrame(
    [
        {'Assay/Protocol': 'Moya CLIA (=CLIApre)', 'Median (IQR)': '433 (98 to >1380)'},
        {'Assay/Protocol': 'Current CLIApost', 'Median (IQR)': '160.02 (30.35 to 514.55)'},
        {'Assay/Protocol': 'Moya RIA', 'Median (IQR)': '85 (5 to 476)'},
        {'Assay/Protocol': 'Current RIA', 'Median (IQR)': '200 (93 to 459)'},
    ]
)


table_5 = pd.DataFrame(
    [
        {'Assay/Protocol': 'Moya CLIA (=CLIApre)', 'Optimal cutpoint': '<=10', 'Sensitivity': '100%', 'Specificity': '100%'},
        {'Assay/Protocol': 'Current CLIApost', 'Optimal cutpoint': '<=4.6', 'Sensitivity': '96.9%', 'Specificity': '93.7%'},
        {'Assay/Protocol': 'Moya RIA', 'Optimal cutpoint': '<=2', 'Sensitivity': '97.2%', 'Specificity': '93.6%'},
        {'Assay/Protocol': 'Current RIA', 'Optimal cutpoint': '<=3.0', 'Sensitivity': '91.7%', 'Specificity': '99.4%'},
    ]
)


table_6 = pd.DataFrame(
    [
        {'Assay/Protocol': 'Current CLIApost', 'Optimal cutpoint': '>=4.83', 'Sensitivity': '94.7%', 'Specificity': '86.5%'},
        {'Assay/Protocol': 'Current RIA', 'Optimal cutpoint': '>=4.0', 'Sensitivity': '99.4%', 'Specificity': '91.7%'},
    ]
)


table_7 = pd.DataFrame(
    [
        {
            'Outcome': 'HA diagnosis (post-ACTH <=2 ug/dL)',
            'Urine test type': 'CLIApost',
            'Cutoff for 100% sensitivity': '<=6.3',
            'Cutoff for 100% specificity': '<=1.98',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA diagnosis (post-ACTH <=2 ug/dL)',
            'Urine test type': 'RIA',
            'Cutoff for 100% sensitivity': '<=26',
            'Cutoff for 100% specificity': '<=0',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)',
            'Urine test type': 'CLIApost',
            'Cutoff for 100% sensitivity': '>=2.01',
            'Cutoff for 100% specificity': '>=73.34',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)',
            'Urine test type': 'RIA',
            'Cutoff for 100% sensitivity': '>=1',
            'Cutoff for 100% specificity': '>=27',
            'Cutoff for 100% accuracy': 'none',
        },
    ]
)


table_8 = pd.DataFrame(
    [
        {
            'Assay/Protocol': 'CLIApost',
            'Optimal cutpoint (nmol/L)': '<=27.59',
            'Sensitivity': '96.9%',
            'Specificity': '93.7%',
            'PPV': '83.8%',
            'NPV': '98.9%',
            'Accuracy': '94.5%',
            'n (pos/neg)': '127 (32/95)',
        },
        {
            'Assay/Protocol': 'RIA',
            'Optimal cutpoint (nmol/L)': '<=55.00',
            'Sensitivity': '100.0%',
            'Specificity': '93.4%',
            'PPV': '52.2%',
            'NPV': '100.0%',
            'Accuracy': '93.9%',
            'n (pos/neg)': '179 (12/167)',
        },
    ]
)

table_9 = pd.DataFrame(
    [
        {
            'Assay/Protocol': 'CLIApost',
            'Optimal cutpoint (nmol/L)': '>=33.11',
            'Sensitivity': '94.7%',
            'Specificity': '86.5%',
            'PPV': '95.5%',
            'NPV': '84.2%',
            'Accuracy': '92.7%',
            'n (pos/neg)': '150 (113/37)',
        },
        {
            'Assay/Protocol': 'RIA',
            'Optimal cutpoint (nmol/L)': '>=58.00',
            'Sensitivity': '93.4%',
            'Specificity': '100.0%',
            'PPV': '100.0%',
            'NPV': '52.2%',
            'Accuracy': '93.9%',
            'n (pos/neg)': '179 (167/12)',
        },
    ]
)

table_10 = pd.DataFrame(
    [
        {
            'Outcome': 'HA diagnosis (post-ACTH <=2 ug/dL)',
            'Urine test type': 'CLIApost',
            'Cutoff for 100% sensitivity': '<=132.43 nmol/L',
            'Cutoff for 100% specificity': '<=10.00 nmol/L',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA diagnosis (post-ACTH <=2 ug/dL)',
            'Urine test type': 'RIA',
            'Cutoff for 100% sensitivity': '<=55.00 nmol/L',
            'Cutoff for 100% specificity': '<=16.00 nmol/L',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)',
            'Urine test type': 'CLIApost',
            'Cutoff for 100% sensitivity': '>=27.59 nmol/L',
            'Cutoff for 100% specificity': '>=2814.18 nmol/L (not clinically useful)',
            'Cutoff for 100% accuracy': 'none',
        },
        {
            'Outcome': 'HA exclusion (baseline >2 ug/dL OR post-ACTH >2 ug/dL)',
            'Urine test type': 'RIA',
            'Cutoff for 100% sensitivity': '>=19.00 nmol/L',
            'Cutoff for 100% specificity': '>=58.00 nmol/L',
            'Cutoff for 100% accuracy': 'none',
        },
    ]
)


table_11 = build_additional_conditions_table(analysis_results_path, 'UCCR')
table_12 = build_additional_conditions_table(urine_results_path, 'Urine cortisol')
table_13 = build_roc_scenarios_table(roc_summary_path)
table_14 = build_overall_interpretation_table(table_2, table_5, table_6, table_8, table_9, table_11, table_12, table_13)


def write_table(sheet, start_row: int, label: str, subtitle: str, df: pd.DataFrame) -> int:
    sheet.cell(row=start_row, column=1, value=label)
    sheet.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    sheet.cell(row=start_row + 1, column=1, value=subtitle)
    sheet.cell(row=start_row + 1, column=1).font = Font(italic=True)

    header_row = start_row + 3
    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=column_name)
        cell.font = Font(bold=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for col_idx, value in enumerate(row.tolist(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    return header_row + len(df) + 2


with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    # Primary narrative + labeled tables in one sheet
    pd.DataFrame({'UCCR Comparison Summary': ['Moya et al. (2022) vs Current Dataset']}).to_excel(
        writer, sheet_name='Comparison_Report', index=False, header=False
    )

    # Also provide each table as dedicated tab for easier reuse.
    table_2.to_excel(writer, sheet_name='Table_2_Study_Comparison', index=False)
    table_3.to_excel(writer, sheet_name='Table_3_UCCR_IQR', index=False)
    table_4.to_excel(writer, sheet_name='Table_4_Urine_Cortisol', index=False)
    table_5.to_excel(writer, sheet_name='Table_5_Diagnosis_Cutoff', index=False)
    table_6.to_excel(writer, sheet_name='Table_6_Exclusion_Cutoff', index=False)
    table_7.to_excel(writer, sheet_name='Table_7_100pct_Cutoffs', index=False)
    table_8.to_excel(writer, sheet_name='Table_8_UrineCort_Diagnosis', index=False)
    table_9.to_excel(writer, sheet_name='Table_9_UrineCort_Exclusion', index=False)
    table_10.to_excel(writer, sheet_name='Table_10_UrineCort_100pct', index=False)
    table_11.to_excel(writer, sheet_name='Table_11_UCCR_Addl_Conds', index=False)
    table_12.to_excel(writer, sheet_name='Table_12_Urine_Addl_Conds', index=False)
    table_13.to_excel(writer, sheet_name='Table_13_ROC_Scenarios', index=False)
    table_14.to_excel(writer, sheet_name='Table_14_Overall_Summary', index=False)

wb = load_workbook(out_path)
ws = wb['Comparison_Report']

ws['A1'].font = Font(bold=True, size=14)
ws['A3'] = 'Table 1. Assay Naming Note'
ws['A3'].font = Font(bold=True, size=12)
ws['A4'] = "In Moya et al. (2022), 'CLIA' is interpreted here as CLIApre."
ws['A5'] = 'In this dataset, urine assay groups are CLIApost and RIA.'

next_row = 7
next_row = write_table(
    ws,
    next_row,
    'Table 2. Study/Analysis Setup Comparison',
    'Comparison of cohort definitions and serum thresholds.',
    table_2,
)
next_row = write_table(
    ws,
    next_row,
    'Table 3. UCCR Medians (IQR) by Assay',
    'HA versus non-HA distributions by assay/protocol.',
    table_3,
)
next_row = write_table(
    ws,
    next_row,
    'Table 4. Urine Cortisol Medians (nmol/L)',
    'Urine cortisol distributions by assay/protocol.',
    table_4,
)
next_row = write_table(
    ws,
    next_row,
    'Table 5. Optimal UCCR Cutpoints for HA Diagnosis',
    'Current workflow diagnosis definition: post-ACTH <=2 ug/dL.',
    table_5,
)
next_row = write_table(
    ws,
    next_row,
    'Table 6. Optimal UCCR Cutpoints for HA Exclusion',
    'Exclusion definition: baseline >2 ug/dL OR post-ACTH >2 ug/dL.',
    table_6,
)
next_row = write_table(
    ws,
    next_row,
    'Table 7. UCCR Cutoffs for 100% Sensitivity, Specificity, and Accuracy',
    'Computed within each urine test type under the 2 ug/dL diagnosis definition.',
    table_7,
)
next_row = write_table(
    ws,
    next_row,
    'Table 8. Optimal Urine Cortisol Cutpoints for HA Diagnosis',
    'Urine cortisol (nmol/L) threshold sweep; HA defined as post-ACTH <=2 ug/dL; criterion: Youden\'s J.',
    table_8,
)
next_row = write_table(
    ws,
    next_row,
    'Table 9. Optimal Urine Cortisol Cutpoints for HA Exclusion',
    'Urine cortisol (nmol/L) threshold sweep; exclusion defined as baseline >2 ug/dL OR post-ACTH >2 ug/dL; criterion: Youden\'s J.',
    table_9,
)
next_row = write_table(
    ws,
    next_row,
    'Table 10. Urine Cortisol Cutoffs for 100% Sensitivity, Specificity, and Accuracy',
    'Computed within each urine test type under the 2 ug/dL diagnosis definition.',
    table_10,
)
next_row = write_table(
    ws,
    next_row,
    'Table 11. UCCR Additional Condition Analyses',
    'Rows excluded before calculation: post-ACTH missing AND basal cortisol below threshold (<1 or <2 ug/dL).',
    table_11,
)
next_row = write_table(
    ws,
    next_row,
    'Table 12. Urine Cortisol Additional Condition Analyses',
    'Rows excluded before calculation: post-ACTH missing AND basal cortisol below threshold (<1 or <2 ug/dL).',
    table_12,
)
next_row = write_table(
    ws,
    next_row,
    'Table 13. ROC Summary Across Base and Additional Conditions',
    'AUC and best-threshold ROC metrics by scenario, analysis, outcome, and urine test type.',
    table_13,
)
next_row = write_table(
    ws,
    next_row,
    'Table 14. Overall Interpretation Summary',
    'Cross-table synthesis of cohort size, assay performance, scenario effects, and ROC discrimination.',
    table_14,
)

ws.cell(row=next_row, column=1, value='Data Sources')
ws.cell(row=next_row, column=1).font = Font(bold=True)
ws.cell(row=next_row + 1, column=1, value='processed/uccr_merged_standardized.csv')
ws.cell(row=next_row + 2, column=1, value='processed/uccr_merged_standardized_full.csv')
ws.cell(row=next_row + 3, column=1, value='processed/uccr_review_pack.xlsx')

# Improve readability
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
    ws.column_dimensions[col].width = 42

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(out_path)

print(f'Wrote: {out_path}')
print('Sheets: Comparison_Report, Table_2_Study_Comparison, Table_3_UCCR_IQR, Table_4_Urine_Cortisol, Table_5_Diagnosis_Cutoff, Table_6_Exclusion_Cutoff, Table_7_100pct_Cutoffs, Table_8_UrineCort_Diagnosis, Table_9_UrineCort_Exclusion, Table_10_UrineCort_100pct, Table_11_UCCR_Addl_Conds, Table_12_Urine_Addl_Conds, Table_13_ROC_Scenarios, Table_14_Overall_Summary')
